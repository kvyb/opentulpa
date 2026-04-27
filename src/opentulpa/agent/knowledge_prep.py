"""Compile uploaded source files into workflow-ready Markdown knowledge packs."""

from __future__ import annotations

import csv
import re
from collections.abc import Iterable
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from opentulpa.agent.file_analysis import extract_uploaded_text

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MAX_WORKBOOK_SHEETS = 12
_MAX_ROWS_PER_SHEET = 500
_MAX_COLS_PER_SHEET = 30
_MAX_CELL_CHARS = 300
_MAX_SOURCE_CHARS = 120_000
_MAX_PACK_CHARS = 260_000
_MAX_STRUCTURE_SHEETS = 40
_MAX_STRUCTURE_SAMPLE_ROWS = 8
_MAX_STRUCTURE_SAMPLE_COLS = 12
_MAX_MATCHES_PER_SHEET = 12
_MAX_TABLE_CANDIDATES_PER_SHEET = 12


def normalize_knowledge_filename(value: Any) -> str:
    raw = str(value or "").strip() or "intake_workflow_knowledge.md"
    if not raw.lower().endswith(".md"):
        raw = f"{raw}.md"
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", raw).strip("._")
    return (safe or "intake_workflow_knowledge.md")[:180]


def normalize_hints(value: Any) -> list[str]:
    if isinstance(value, str):
        candidates: Iterable[Any] = re.split(r"[\n,;]+", value)
    elif isinstance(value, list):
        candidates = value
    else:
        candidates = []
    out: list[str] = []
    seen: set[str] = set()
    for item in candidates:
        text = str(item or "").strip()
        folded = text.casefold()
        if not text or folded in seen:
            continue
        seen.add(folded)
        out.append(text[:120])
    return out[:20]


def normalize_selected_sections(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    sections: list[dict[str, Any]] = []
    for item in value:
        if isinstance(item, str):
            sheet_name = item.strip()
            raw: dict[str, Any] = {"sheet_name": sheet_name} if sheet_name else {}
        elif isinstance(item, dict):
            raw = item
        else:
            continue
        file_id = str(raw.get("file_id", "") or "").strip()
        sheet_name = str(raw.get("sheet_name", raw.get("sheet", "")) or "").strip()
        if not sheet_name:
            continue
        section: dict[str, Any] = {"sheet_name": sheet_name[:120]}
        if file_id:
            section["file_id"] = file_id[:120]
        row_start = _safe_positive_int(raw.get("row_start", raw.get("start_row")))
        row_end = _safe_positive_int(raw.get("row_end", raw.get("end_row")))
        if row_start is not None:
            section["row_start"] = row_start
        if row_end is not None:
            section["row_end"] = row_end
        sections.append(section)
    return sections[:60]


def inspect_uploaded_file_structure(
    *,
    raw_bytes: bytes,
    filename: str | None,
    mime_type: str | None,
    search_terms: Any = None,
) -> dict[str, Any]:
    """Return a compact structural map so an agent can choose useful regions."""
    terms = normalize_hints(search_terms)
    safe_filename = str(filename or "file.bin").strip() or "file.bin"
    safe_mime = str(mime_type or "").strip().lower()
    name = safe_filename.lower()
    if safe_mime == _XLSX_MIME or name.endswith(".xlsx"):
        return _inspect_xlsx_structure(
            raw_bytes=raw_bytes,
            filename=safe_filename,
            search_terms=terms,
        )
    if name.endswith(".csv") or safe_mime in {"text/csv", "application/csv"}:
        return _inspect_delimited_text(raw_bytes=raw_bytes, filename=safe_filename, format_name="csv")

    extracted = extract_uploaded_text(
        raw_bytes=raw_bytes,
        filename=safe_filename,
        mime_type=safe_mime,
        max_chars=16_000,
    )
    return {
        "filename": safe_filename,
        "mime_type": safe_mime,
        "format": "text" if extracted else "unknown",
        "warnings": [] if extracted else [f"no extractable text for {safe_filename}"],
        "structure": {
            "line_count": len(str(extracted or "").splitlines()),
            "preview": str(extracted or "")[:4000],
        },
    }


def build_intake_knowledge_markdown(
    *,
    sources: list[dict[str, Any]],
    workflow_goal: Any = "",
    include_hints: Any = None,
    selected_sections: Any = None,
) -> dict[str, Any]:
    """Return a compiled Markdown pack from file-vault source records and bytes."""
    hints = normalize_hints(include_hints)
    goal = str(workflow_goal or "").strip()
    warnings: list[str] = []
    matched_sections: list[str] = []
    requires_selection = False
    source_file_ids: list[str] = []
    sections = normalize_selected_sections(selected_sections)
    parts: list[str] = ["# Workflow Knowledge Pack\n"]
    if goal:
        parts.append(f"Workflow goal: {_clean_inline(goal, limit=600)}\n")
    if hints:
        parts.append("Relevant scope hints: " + ", ".join(hints) + "\n")
    parts.append(
        "Runtime use: answer and extract workflow facts from this pack first. "
        "If a customer asks outside this scoped knowledge, ask a clarifying question or escalate instead of guessing.\n"
    )

    for index, source in enumerate(sources, start=1):
        raw_record = source.get("record")
        record: dict[str, Any] = raw_record if isinstance(raw_record, dict) else {}
        raw_bytes = bytes(source.get("raw_bytes") or b"")
        file_id = str(record.get("id", "") or "").strip()
        filename = str(record.get("original_filename", "") or "file.bin").strip()
        mime_type = str(record.get("mime_type", "") or "").strip()
        if file_id:
            source_file_ids.append(file_id)
        converted = convert_uploaded_file_to_markdown(
            raw_bytes=raw_bytes,
            filename=filename,
            mime_type=mime_type,
            include_hints=hints,
            selected_sections=sections,
            file_id=file_id,
        )
        warnings.extend(converted.get("warnings") or [])
        matched_sections.extend(str(item) for item in converted.get("matched_sections") or [])
        requires_selection = requires_selection or bool(converted.get("requires_selection"))
        parts.append(
            "\n\n"
            f"## Source {index}: {_clean_inline(filename, limit=180)}\n\n"
            f"- file_id: `{file_id}`\n"
            f"- mime_type: `{mime_type or 'unknown'}`\n\n"
            f"{str(converted.get('markdown', '')).strip()}"
        )

    markdown = "\n".join(parts).strip()
    if len(markdown) > _MAX_PACK_CHARS:
        warnings.append(f"compiled knowledge pack truncated to {_MAX_PACK_CHARS} characters")
        markdown = markdown[:_MAX_PACK_CHARS].rstrip() + "\n\n[truncated]\n"

    return {
        "markdown": markdown,
        "source_file_ids": source_file_ids,
        "warnings": warnings,
        "matched_sections": _unique_texts(matched_sections),
        "requires_selection": requires_selection,
    }


def convert_uploaded_file_to_markdown(
    *,
    raw_bytes: bytes,
    filename: str | None,
    mime_type: str | None,
    include_hints: Any = None,
    selected_sections: Any = None,
    file_id: str | None = None,
) -> dict[str, Any]:
    hints = normalize_hints(include_hints)
    sections = normalize_selected_sections(selected_sections)
    safe_filename = str(filename or "file.bin").strip() or "file.bin"
    safe_mime = str(mime_type or "").strip().lower()
    name = safe_filename.lower()
    if safe_mime == _XLSX_MIME or name.endswith(".xlsx"):
        return _xlsx_to_markdown(
            raw_bytes=raw_bytes,
            filename=safe_filename,
            include_hints=hints,
            selected_sections=sections,
            file_id=str(file_id or "").strip(),
        )
    if name.endswith(".csv") or safe_mime in {"text/csv", "application/csv"}:
        return _csv_to_markdown(raw_bytes=raw_bytes, filename=safe_filename)

    extracted = extract_uploaded_text(
        raw_bytes=raw_bytes,
        filename=safe_filename,
        mime_type=safe_mime,
        max_chars=_MAX_SOURCE_CHARS,
    )
    if extracted:
        return {
            "markdown": "```text\n" + _fence_safe(extracted) + "\n```",
            "warnings": [],
            "matched_sections": [],
        }
    return {
        "markdown": "_No extractable text was available for this source._",
        "warnings": [f"no extractable text for {safe_filename}"],
        "matched_sections": [],
    }


def _xlsx_to_markdown(
    *,
    raw_bytes: bytes,
    filename: str,
    include_hints: list[str],
    selected_sections: list[dict[str, Any]],
    file_id: str,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
        from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
    except Exception as exc:  # pragma: no cover - exercised only without dependency
        return {
            "markdown": "_XLSX parsing is unavailable because openpyxl is not installed._",
            "warnings": [f"xlsx parser unavailable for {filename}: {exc}"],
            "matched_sections": [],
        }

    warnings: list[str] = []
    try:
        workbook = load_workbook(BytesIO(raw_bytes), data_only=True, read_only=False)
    except Exception as exc:
        return {
            "markdown": "_XLSX parsing failed._",
            "warnings": [f"xlsx parsing failed for {filename}: {exc}"],
            "matched_sections": [],
        }

    sheet_names = list(workbook.sheetnames)
    selected: list[tuple[str, str, int | None, int | None]] = []
    matched_by_name = {name.casefold(): name for name in sheet_names}
    for section in selected_sections:
        section_file_id = str(section.get("file_id", "") or "").strip()
        if section_file_id and file_id and section_file_id != file_id:
            continue
        raw_sheet = str(section.get("sheet_name", "") or "").strip()
        sheet_name = matched_by_name.get(raw_sheet.casefold())
        if not sheet_name:
            warnings.append(f"selected sheet not found in {filename}: {raw_sheet}")
            continue
        selected.append(
            (
                sheet_name,
                "selected_section",
                _safe_positive_int(section.get("row_start")),
                _safe_positive_int(section.get("row_end")),
            )
        )
    if not selected:
        for sheet_name in sheet_names:
            if _matches_any_hint(sheet_name, include_hints):
                selected.append((sheet_name, "sheet_name", None, None))
    if include_hints and not selected:
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            if _worksheet_has_hint(sheet, include_hints):
                selected.append((sheet_name, "cell_match", None, None))
    if len(selected) > _MAX_WORKBOOK_SHEETS:
        warnings.append(f"selected sheet count truncated to {_MAX_WORKBOOK_SHEETS}")
        selected = selected[:_MAX_WORKBOOK_SHEETS]

    parts = [
        f"Workbook sheets: {', '.join(_clean_inline(name, limit=80) for name in sheet_names)}"
    ]
    matched_sections = [_selected_section_ref(filename, name, start, end) for name, _, start, end in selected]
    if not selected:
        if include_hints:
            warnings.append(f"no workbook sheets matched hints: {', '.join(include_hints)}")
            parts.append("\n_No workbook sheets matched the requested scope hints._")
        else:
            warnings.append("xlsx selection required: inspect the workbook and pass selected_sections")
            parts.append("\n_No workbook sheets were selected. Inspect the workbook structure first, then prepare selected sheets or row ranges._")
        return {
            "markdown": "\n\n".join(parts),
            "warnings": warnings,
            "matched_sections": matched_sections,
            "requires_selection": True,
        }

    for sheet_name, reason, row_start, row_end in selected:
        sheet = workbook[sheet_name]
        parts.append(
            _worksheet_to_markdown(
                sheet=sheet,
                sheet_name=sheet_name,
                selection_reason=reason,
                include_hints=include_hints,
                row_start=row_start,
                row_end=row_end,
                get_column_letter=get_column_letter,
            )
        )

    return {
        "markdown": "\n\n".join(part for part in parts if str(part).strip()),
        "warnings": warnings,
        "matched_sections": matched_sections,
        "requires_selection": False,
    }


def _worksheet_to_markdown(
    *,
    sheet: Any,
    sheet_name: str,
    selection_reason: str,
    include_hints: list[str],
    row_start: int | None,
    row_end: int | None,
    get_column_letter: Any,
) -> str:
    merged_values = _merged_cell_values(sheet)
    nonempty_rows: list[int] = []
    row_matches: set[int] = set()
    for row_index in range(1, int(sheet.max_row or 0) + 1):
        values = [
            _cell_text(_cell_value(sheet, row_index, col_index, merged_values))
            for col_index in range(1, int(sheet.max_column or 0) + 1)
        ]
        if any(values):
            nonempty_rows.append(row_index)
        if include_hints and any(_matches_any_hint(value, include_hints) for value in values):
            row_matches.add(row_index)

    if row_start is not None or row_end is not None:
        first = row_start or 1
        last = row_end or int(sheet.max_row or first)
        if last < first:
            first, last = last, first
        row_numbers = [row for row in nonempty_rows if first <= row <= last]
    elif selection_reason == "cell_match" and row_matches:
        row_numbers = [
            row
            for row in nonempty_rows
            if any(abs(row - matched) <= 2 for matched in row_matches)
        ]
    else:
        row_numbers = list(nonempty_rows)
    truncated = False
    if len(row_numbers) > _MAX_ROWS_PER_SHEET:
        row_numbers = row_numbers[:_MAX_ROWS_PER_SHEET]
        truncated = True

    used_cols: list[int] = []
    for col_index in range(1, int(sheet.max_column or 0) + 1):
        if any(_cell_text(_cell_value(sheet, row, col_index, merged_values)) for row in row_numbers):
            used_cols.append(col_index)
    if len(used_cols) > _MAX_COLS_PER_SHEET:
        used_cols = used_cols[:_MAX_COLS_PER_SHEET]
        truncated = True

    if not row_numbers or not used_cols:
        return f"### Sheet: {sheet_name}\n\n_No non-empty rows found._"

    header = ["source_ref", *[str(get_column_letter(col)) for col in used_cols]]
    rows: list[list[str]] = []
    for row_index in row_numbers:
        rows.append(
            [
                f"{sheet_name}!{row_index}",
                *[
                    _cell_text(_cell_value(sheet, row_index, col_index, merged_values))
                    for col_index in used_cols
                ],
            ]
        )
    suffix = "\n\n_Note: rows or columns were truncated for prompt safety._" if truncated else ""
    title = f"### Sheet: {sheet_name}"
    if row_start is not None or row_end is not None:
        title += f" rows {row_start or 1}-{row_end or int(sheet.max_row or 0)}"
    return f"{title}\n\n{_markdown_table(header, rows)}{suffix}"


def _inspect_xlsx_structure(
    *,
    raw_bytes: bytes,
    filename: str,
    search_terms: list[str],
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except Exception as exc:  # pragma: no cover - exercised only without dependency
        return {
            "filename": filename,
            "mime_type": _XLSX_MIME,
            "format": "xlsx",
            "warnings": [f"xlsx parser unavailable for {filename}: {exc}"],
            "structure": {"sheets": []},
        }

    warnings: list[str] = []
    try:
        workbook = load_workbook(BytesIO(raw_bytes), data_only=True, read_only=False)
    except Exception as exc:
        return {
            "filename": filename,
            "mime_type": _XLSX_MIME,
            "format": "xlsx",
            "warnings": [f"xlsx parsing failed for {filename}: {exc}"],
            "structure": {"sheets": []},
        }

    sheets: list[dict[str, Any]] = []
    sheet_names = list(workbook.sheetnames)
    if len(sheet_names) > _MAX_STRUCTURE_SHEETS:
        warnings.append(f"sheet inventory truncated to {_MAX_STRUCTURE_SHEETS}")
    for index, sheet_name in enumerate(sheet_names[:_MAX_STRUCTURE_SHEETS], start=1):
        sheet = workbook[sheet_name]
        merged_values = _merged_cell_values(sheet)
        nonempty_rows: list[tuple[int, list[str]]] = []
        matches: list[dict[str, Any]] = []
        matched_terms = [
            term for term in search_terms if _matches_any_hint(sheet_name, [term])
        ]
        for row_index in range(1, int(sheet.max_row or 0) + 1):
            values = [
                _cell_text(_cell_value(sheet, row_index, col_index, merged_values))
                for col_index in range(1, min(int(sheet.max_column or 0), _MAX_STRUCTURE_SAMPLE_COLS) + 1)
            ]
            full_values = values
            if int(sheet.max_column or 0) > _MAX_STRUCTURE_SAMPLE_COLS and search_terms:
                full_values = [
                    _cell_text(_cell_value(sheet, row_index, col_index, merged_values))
                    for col_index in range(1, int(sheet.max_column or 0) + 1)
                ]
            if any(values) or any(full_values):
                nonempty_rows.append((row_index, values))
            if search_terms and len(matches) < _MAX_MATCHES_PER_SHEET:
                row_text = " ".join(full_values)
                matched_term = next((term for term in search_terms if _matches_any_hint(row_text, [term])), "")
                if matched_term:
                    matches.append(
                        {
                            "term": matched_term,
                            "source_ref": f"{sheet_name}!{row_index}",
                            "row": row_index,
                            "values": _trim_row_values(full_values),
                        }
                    )
        include_details = not search_terms or bool(matched_terms)
        sample_rows = [
            {"source_ref": f"{sheet_name}!{row}", "row": row, "values": _trim_row_values(values)}
            for row, values in nonempty_rows[:_MAX_STRUCTURE_SAMPLE_ROWS]
        ] if include_details else []
        table_candidates = _table_candidates(sheet_name, nonempty_rows) if include_details else []
        visible_matches = matches if include_details else matches[:3]
        omitted_reason = ""
        if not include_details:
            omitted_reason = (
                "sample rows and table candidates omitted because this sheet did not match "
                "search_terms by sheet name; see matches for relevant cell hits"
            )
        sheets.append(
            {
                "index": index,
                "name": sheet_name,
                "matched_terms": matched_terms,
                "max_row": int(sheet.max_row or 0),
                "max_column": int(sheet.max_column or 0),
                "nonempty_rows": len(nonempty_rows),
                "sample_rows": sample_rows,
                "table_candidates": table_candidates,
                "matches": visible_matches,
                **({"omitted_detail_reason": omitted_reason} if omitted_reason else {}),
            }
        )

    return {
        "filename": filename,
        "mime_type": _XLSX_MIME,
        "format": "xlsx",
        "warnings": warnings,
        "structure": {
            "sheets": sheets,
            "selection_format": {
                "file_id": "optional source file id when preparing multiple files",
                "sheet_name": "exact sheet name from this inventory",
                "row_start": "optional 1-based first row",
                "row_end": "optional 1-based last row",
            },
        },
    }


def _inspect_delimited_text(
    *,
    raw_bytes: bytes,
    filename: str,
    format_name: str,
) -> dict[str, Any]:
    text = raw_bytes.decode("utf-8", errors="replace")
    rows = list(csv.reader(StringIO(text)))[:_MAX_STRUCTURE_SAMPLE_ROWS]
    return {
        "filename": filename,
        "mime_type": "text/csv",
        "format": format_name,
        "warnings": [],
        "structure": {
            "sample_rows": [
                {"row": index, "values": _trim_row_values(row)}
                for index, row in enumerate(rows, start=1)
            ],
        },
    }


def _table_candidates(sheet_name: str, nonempty_rows: list[tuple[int, list[str]]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    current: list[tuple[int, list[str]]] = []
    previous_row = 0
    for row_number, values in nonempty_rows:
        if current and row_number > previous_row + 1:
            candidates.append(_table_candidate(sheet_name, current))
            current = []
        current.append((row_number, values))
        previous_row = row_number
    if current:
        candidates.append(_table_candidate(sheet_name, current))
    return candidates[:_MAX_TABLE_CANDIDATES_PER_SHEET]


def _table_candidate(sheet_name: str, rows: list[tuple[int, list[str]]]) -> dict[str, Any]:
    first = rows[0][0]
    last = rows[-1][0]
    sample = rows[: min(len(rows), 3)]
    return {
        "sheet_name": sheet_name,
        "row_start": first,
        "row_end": last,
        "nonempty_rows": len(rows),
        "sample_rows": [
            {"source_ref": f"{sheet_name}!{row}", "row": row, "values": _trim_row_values(values)}
            for row, values in sample
        ],
    }


def _csv_to_markdown(*, raw_bytes: bytes, filename: str) -> dict[str, Any]:
    text = raw_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = list(reader)[:_MAX_ROWS_PER_SHEET]
    if not rows:
        return {"markdown": "_CSV file was empty._", "warnings": [], "matched_sections": []}
    max_cols = min(max(len(row) for row in rows), _MAX_COLS_PER_SHEET)
    header = [f"col_{index}" for index in range(1, max_cols + 1)]
    body = [[_clean_inline(cell, limit=_MAX_CELL_CHARS) for cell in row[:max_cols]] for row in rows]
    warnings = []
    if len(rows) >= _MAX_ROWS_PER_SHEET:
        warnings.append(f"csv rows truncated for {filename}")
    return {
        "markdown": _markdown_table(header, body),
        "warnings": warnings,
        "matched_sections": [],
    }


def _merged_cell_values(sheet: Any) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    ranges = getattr(getattr(sheet, "merged_cells", None), "ranges", []) or []
    for merged in ranges:
        top_value = sheet.cell(int(merged.min_row), int(merged.min_col)).value
        if top_value is None:
            continue
        for row_index in range(int(merged.min_row), int(merged.max_row) + 1):
            for col_index in range(int(merged.min_col), int(merged.max_col) + 1):
                values[(row_index, col_index)] = top_value
    return values


def _cell_value(sheet: Any, row_index: int, col_index: int, merged_values: dict[tuple[int, int], Any]) -> Any:
    value = sheet.cell(row_index, col_index).value
    if value is None:
        return merged_values.get((row_index, col_index))
    return value


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean_inline(value, limit=_MAX_CELL_CHARS)


def _clean_inline(value: Any, *, limit: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text[:limit]


def _matches_any_hint(value: Any, hints: list[str]) -> bool:
    if not hints:
        return False
    text = str(value or "").casefold()
    return any(hint.casefold() in text or text in hint.casefold() for hint in hints if hint)


def _worksheet_has_hint(sheet: Any, hints: list[str]) -> bool:
    if not hints:
        return False
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if _matches_any_hint(value, hints):
                return True
    return False


def _markdown_table(header: list[str], rows: list[list[str]]) -> str:
    safe_header = [_escape_table_cell(cell) for cell in header]
    out = [
        "| " + " | ".join(safe_header) + " |",
        "| " + " | ".join("---" for _ in safe_header) + " |",
    ]
    for row in rows:
        cells = [_escape_table_cell(cell) for cell in row]
        if len(cells) < len(safe_header):
            cells.extend([""] * (len(safe_header) - len(cells)))
        out.append("| " + " | ".join(cells[: len(safe_header)]) + " |")
    return "\n".join(out)


def _escape_table_cell(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ").strip()


def _fence_safe(value: str) -> str:
    return str(value or "").replace("```", "`\u200b``")


def _unique_texts(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        folded = text.casefold()
        if not text or folded in seen:
            continue
        seen.add(folded)
        out.append(text)
    return out


def _safe_positive_int(value: Any) -> int | None:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    if parsed <= 0:
        return None
    return parsed


def _selected_section_ref(
    filename: str,
    sheet_name: str,
    row_start: int | None,
    row_end: int | None,
) -> str:
    base = f"{filename}!{sheet_name}"
    if row_start is None and row_end is None:
        return base
    return f"{base}:{row_start or 1}-{row_end or ''}"


def _trim_row_values(values: list[Any]) -> list[str]:
    return [
        _clean_inline(value, limit=_MAX_CELL_CHARS)
        for value in values[:_MAX_STRUCTURE_SAMPLE_COLS]
    ]
