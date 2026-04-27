from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from opentulpa.agent.file_analysis import extract_uploaded_text
from opentulpa.agent.knowledge_prep import (
    build_intake_knowledge_markdown,
    inspect_uploaded_file_structure,
)
from opentulpa.api.app import create_app
from opentulpa.context.file_vault import FileVaultService


class _DisabledComposio:
    enabled = False

    def status(self) -> dict[str, object]:
        return {"ok": True, "enabled": False}


def _autospa_workbook_bytes() -> bytes:
    workbook = Workbook()
    wash = workbook.active
    wash.title = "Мойка"
    wash.append(["Услуга", "C-Class", "SUV"])
    wash.append(["2х-фазная мойка кузова", 1000, 1200])
    tire = workbook.create_sheet("Шиномонтаж")
    tire.append(
        [
            "Размерность дисков",
            "Седан",
            "Внедорожник / кросовер + низкий профиль",
        ]
    )
    tire.append(["Комплект 19R", 3000, 4000])
    ppf = workbook.create_sheet("PPF")
    ppf.append(["Пакет", "Цена"])
    ppf.append(["Передняя часть", 50000])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_xlsx_knowledge_pack_scopes_selected_sheets() -> None:
    prepared = build_intake_knowledge_markdown(
        sources=[
            {
                "record": {
                    "id": "file_1",
                    "original_filename": "autospa.xlsx",
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                "raw_bytes": _autospa_workbook_bytes(),
            }
        ],
        workflow_goal="Handle мойка and шиномонтаж bookings.",
        include_hints=["Шиномонтаж"],
    )

    markdown = prepared["markdown"]
    assert "Workflow Knowledge Pack" in markdown
    assert "### Sheet: Шиномонтаж" in markdown
    assert "Комплект 19R" in markdown
    assert "4000" in markdown
    assert "### Sheet: Мойка" not in markdown
    assert "autospa.xlsx!Шиномонтаж" in prepared["matched_sections"]


def test_xlsx_inspection_returns_structure_and_search_matches() -> None:
    inspected = inspect_uploaded_file_structure(
        raw_bytes=_autospa_workbook_bytes(),
        filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        search_terms=["Комплект", "Шиномонтаж"],
    )

    sheets = inspected["structure"]["sheets"]
    assert [sheet["name"] for sheet in sheets] == ["Мойка", "Шиномонтаж", "PPF"]
    tire_sheet = sheets[1]
    assert tire_sheet["matched_terms"] == ["Шиномонтаж"]
    assert tire_sheet["sample_rows"][0]["source_ref"] == "Шиномонтаж!1"
    assert tire_sheet["table_candidates"][0]["row_start"] == 1
    assert tire_sheet["matches"][0]["source_ref"] == "Шиномонтаж!2"


def test_xlsx_upload_text_extraction_returns_workbook_preview() -> None:
    extracted = extract_uploaded_text(
        raw_bytes=_autospa_workbook_bytes(),
        filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    assert "# Sheet 1: Мойка" in extracted
    assert "2х-фазная мойка кузова" in extracted
    assert "# Sheet 2: Шиномонтаж" in extracted
    assert "Комплект 19R" in extracted


def test_xlsx_knowledge_pack_accepts_explicit_selected_sections_without_hints() -> None:
    prepared = build_intake_knowledge_markdown(
        sources=[
            {
                "record": {
                    "id": "file_1",
                    "original_filename": "autospa.xlsx",
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                "raw_bytes": _autospa_workbook_bytes(),
            }
        ],
        workflow_goal="Handle only source sections chosen during setup.",
        selected_sections=[
            {
                "file_id": "file_1",
                "sheet_name": "Мойка",
                "row_start": 1,
                "row_end": 2,
            }
        ],
    )

    markdown = prepared["markdown"]
    assert prepared["requires_selection"] is False
    assert "### Sheet: Мойка rows 1-2" in markdown
    assert "2х-фазная мойка кузова" in markdown
    assert "Комплект 19R" not in markdown


def test_xlsx_knowledge_pack_requires_selection_without_hints() -> None:
    prepared = build_intake_knowledge_markdown(
        sources=[
            {
                "record": {
                    "id": "file_1",
                    "original_filename": "autospa.xlsx",
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                "raw_bytes": _autospa_workbook_bytes(),
            }
        ],
        workflow_goal="Handle bookings.",
    )

    assert prepared["requires_selection"] is True
    assert "inspect the workbook" in prepared["markdown"].lower()


def test_prepare_intake_knowledge_route_creates_markdown_file(tmp_path: Path) -> None:
    file_vault = FileVaultService(
        root_dir=tmp_path / "file_vault",
        db_path=tmp_path / "file_vault.db",
    )
    source = file_vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        caption=None,
        raw_bytes=_autospa_workbook_bytes(),
    )
    app = create_app(
        file_vault_service=file_vault,
        composio_service=_DisabledComposio(),
    )

    with TestClient(app) as client:
        response = client.post(
            "/internal/files/prepare_intake_knowledge",
            json={
                "customer_id": "telegram_123",
                "file_ids": [source["id"]],
                "selected_sections": [
                    {
                        "file_id": source["id"],
                        "sheet_name": "Мойка",
                        "row_start": 1,
                        "row_end": 2,
                    },
                    {
                        "file_id": source["id"],
                        "sheet_name": "Шиномонтаж",
                        "row_start": 1,
                        "row_end": 2,
                    },
                ],
                "workflow_goal": "Book car wash and tire fitting requests.",
                "output_name": "autospa_intake_knowledge.md",
            },
        )

    assert response.status_code == 200
    payload = response.json()
    knowledge_file_id = payload["knowledge_file_id"]
    assert knowledge_file_id
    prepared = file_vault.get_file("telegram_123", knowledge_file_id)
    assert prepared is not None
    assert prepared["kind"] == "workflow_knowledge"
    assert prepared["original_filename"] == "autospa_intake_knowledge.md"
    assert "2х-фазная мойка кузова" in prepared["text_excerpt"]
    assert "Комплект 19R" in prepared["text_excerpt"]


def test_prepare_intake_knowledge_route_requires_xlsx_selection(tmp_path: Path) -> None:
    file_vault = FileVaultService(
        root_dir=tmp_path / "file_vault",
        db_path=tmp_path / "file_vault.db",
    )
    source = file_vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        caption=None,
        raw_bytes=_autospa_workbook_bytes(),
    )
    app = create_app(
        file_vault_service=file_vault,
        composio_service=_DisabledComposio(),
    )

    with TestClient(app) as client:
        response = client.post(
            "/internal/files/prepare_intake_knowledge",
            json={
                "customer_id": "telegram_123",
                "file_ids": [source["id"]],
                "workflow_goal": "Book car wash and tire fitting requests.",
            },
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["needs_selection"] is True
    assert "selected_sections" in payload["detail"]


def test_inspect_structure_route_returns_workbook_inventory(tmp_path: Path) -> None:
    file_vault = FileVaultService(
        root_dir=tmp_path / "file_vault",
        db_path=tmp_path / "file_vault.db",
    )
    source = file_vault.ingest_file(
        customer_id="telegram_123",
        chat_id=None,
        kind="document",
        telegram_file_id=None,
        original_filename="autospa.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        caption=None,
        raw_bytes=_autospa_workbook_bytes(),
    )
    app = create_app(
        file_vault_service=file_vault,
        composio_service=_DisabledComposio(),
    )

    with TestClient(app) as client:
        response = client.post(
            "/internal/files/inspect_structure",
            json={
                "customer_id": "telegram_123",
                "file_id": source["id"],
                "search_terms": ["низкий профиль", "Шиномонтаж"],
            },
        )

    assert response.status_code == 200
    inspection = response.json()["inspection"]
    assert inspection["format"] == "xlsx"
    tire_sheet = inspection["structure"]["sheets"][1]
    assert tire_sheet["name"] == "Шиномонтаж"
    assert tire_sheet["matched_terms"] == ["Шиномонтаж"]
    assert tire_sheet["matches"][0]["source_ref"] == "Шиномонтаж!1"
